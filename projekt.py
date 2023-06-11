import datetime
import os
import regex as re
import shutil
from openpyxl import load_workbook


print('Podaj lokalizację.')
path = input() #r'C:\Users\48690\PycharmProjects\python\ZUS-przelewy-pythonProject\przelewy ZUS\17.04.2023'
lokalizacja_plikow = path

shutil.rmtree('data - nowy folder')

os.mkdir('data - nowy folder')
print('Utworzono nowy folder o nazwie: "data - nowy folder\n"')

lista_plikow_w_folderze = os.listdir(path) #przeczytane pliki w folderze

orginalna_sciezka = path
docelowa_sciezka = r'C:\Users\48690\PycharmProjects\python\ZUS-przelewy-pythonProject\przelewy ZUS\data - nowy folder'


for element in lista_plikow_w_folderze:
    nazwa_pliku = orginalna_sciezka + '\\' + element
    docelowa_nazwa_pliku = docelowa_sciezka + '\\' + element
    shutil.copy(nazwa_pliku, docelowa_nazwa_pliku, follow_symlinks=True)
print('Pliki zostały skopiowane.\n')

# DO TEGO MIEJSCA JEST STWORZONY NOWY FODER Z SKOPIOWANYMI PLIKAMI Z POPRZEDNIEGO MIESIACA

month_before_name = '0'+ str(datetime.date.today().month)
month_current_name = '0'+ str(datetime.date.today().month + 1 )

folder = docelowa_sciezka
objs = os.listdir(folder)

for src in objs:
   full_src = os.path.join(folder, src)
   if os.path.isfile(full_src):
       dst = src.replace(month_before_name, month_current_name)
       if src != dst:
           full_dst = os.path.join(folder, dst)
           os.rename(full_src, full_dst)
print('Nazwy plików zostały zmienione.\n')

# DO TEGO MIEJSCA JEST SA ZMIENIONE NAZWY PLIKÓW ZGODNIE Z MIESIACEM

lista_plikow_w_nowym_folderze = os.listdir(docelowa_sciezka) #przeczytane pliki w folderze o nazwie 'data - nowy folder'
print(lista_plikow_w_nowym_folderze) #LISTA PLIKÓW W FOLDERZE

myregex = '^[0-9]+,[0-9]{8},[0-9]+,[0-9]+,[0-9],"[0-9]+","[0-9]+","[^"]+","[^"]+",'
regex_kwota = r'^[0-9]{3},[0-9]{8},\K[0-9]+'
regex_pusta_kwota = '^[0-9]{3},[0-9]{8},'
lista_udało_sie = []
nie_udalo_sie = []


wb = load_workbook("MATRYCA DO WYSYŁKI PRZELEWÓW.xlsx")
sheet = wb.active

max_row = sheet.max_row
max_column = sheet.max_column


for element in lista_plikow_w_nowym_folderze: #CZYTANIE PLIKÓW Z NOWEGO FOLDERU
    docelowa_sciezka = os.path.join(docelowa_sciezka)
    print(docelowa_sciezka + '\\' + element)


    open_file = open(docelowa_sciezka + '\\' + element, encoding="ANSI")
    zawartosc_txt = open_file.read()

    print('\n')
    print(zawartosc_txt)
    x = re.search(myregex, zawartosc_txt)
    if x:
        lista_udało_sie.append(element[0:4])
        print('\n Udało się, plik: ' + element)
        for i in range(1, max_row + 1):
            if str(sheet.cell(row=i, column=1).value) == element[0:4]:
                print('\n\n')
                y = re.search(regex_kwota,zawartosc_txt)
                if y:
                    f = open(docelowa_sciezka + '\\' + element, "w", encoding="ANSI")
                    nowa_zawartosc = re.sub(regex_kwota, str(sheet.cell(row=i, column=6).value), zawartosc_txt)
                    f.write(nowa_zawartosc)
                    f.close()
    else:
        nie_udalo_sie.append(element[0:4])
        print("NIE UDAŁO SIĘ \n" + element)
print(lista_udało_sie)
print(nie_udalo_sie)


# wb = Workbook("MATRYCA DO WYSYŁKI PRZELEWÓW.xlsx")
# collection = wb.worksheets()
# collectionCount = collection.getCount()
#
# for worksheetIndex in range(collectionCount):
#     worksheet = collection.get(worksheetIndex)
#     print("worksheet: " + str(worksheetIndex))









